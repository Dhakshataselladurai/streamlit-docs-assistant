"""
Run Evaluation Set Through the RAG Chatbot
--------------------------------------------
Reads streamlit_rag_eval_set.xlsx, runs every question through your
ask_docs() function, and writes the chatbot's answer into the
"Chatbot Answer" column automatically.

You still need to manually fill in the "Score" column afterward by
reading each answer and comparing it to the "Expected Answer" column.

Setup:
    1. Place this file in your project folder (same folder as rag.py).
    2. Place streamlit_rag_eval_set.xlsx in the same folder too.
    3. Run: python run_eval.py

Output:
    Creates a new file, streamlit_rag_eval_results.xlsx, so your
    original blank template is preserved untouched.
"""

import time
from openpyxl import load_workbook
from rag import ask_docs

INPUT_FILE = "streamlit_rag_eval_set.xlsx"
OUTPUT_FILE = "streamlit_rag_eval_results.xlsx"

# Column positions in the spreadsheet (1-indexed, matching the template)
COL_QUESTION = 3        # C: Question
COL_CHATBOT_ANSWER = 7  # G: Chatbot Answer


def main():
    print(f"Loading {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)
    sheet = wb.active

    total_rows = sheet.max_row - 1  # minus header row
    print(f"Found {total_rows} questions. Running through ask_docs()...\n")

    for row_num in range(2, sheet.max_row + 1):  # row 1 is the header
        question_cell = sheet.cell(row=row_num, column=COL_QUESTION)
        question = question_cell.value

        if not question:
            continue  # skip any blank rows

        print(f"[{row_num - 1}/{total_rows}] Asking: {question}")

        try:
            answer, sources, github_issues = ask_docs(question)
        except Exception as e:
            answer = f"ERROR while generating answer: {e}"
            print(f"  -> Failed: {e}")

        sheet.cell(row=row_num, column=COL_CHATBOT_ANSWER, value=answer)

        # Small delay to avoid hammering the Gemini/GitHub APIs back-to-back
        time.sleep(1)

    wb.save(OUTPUT_FILE)
    print(f"\nDone. Results saved to {OUTPUT_FILE}")
    print("Open it, read each Chatbot Answer next to the Expected Answer,")
    print("and fill in the Score column (Correct / Partial / Wrong).")


if __name__ == "__main__":
    main()